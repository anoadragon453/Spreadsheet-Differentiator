# -*- coding: utf-8 -*-

# Spreadsheet differentiator
# Find differences in spreadsheets and fucking highlight that shit
# Created by Andrew Morgan (2015)

# A-Q, 1-666

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Font to change cells to
newFont = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF0000')

#load workbook
wb = load_workbook('ss.xlsm')
print "Found sheets: " + str(wb.get_sheet_names())

# Grab and store sheets
sheetOrig = wb['Sheet1']
sheetChanged = wb['Sheet2']

# Loop through both sheets and check for differences
letterRange = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']

for indexChangedX in letterRange:
	for indexChangedY in xrange(1,667):
		# Get cell in changed array to check
		changedContentToCheck = sheetChanged[indexChangedX + str(indexChangedY)].value

		# Check if there's a match anywhere
		match = False
		for indexOrigX in letterRange:
			for indexOrigY in xrange(1,667):
				print "Checking Index " + str(indexOrigX) + "," + str(indexOrigY) + " against changed index " + str(indexChangedX) + "," + str(indexChangedY)

				# Get cell in original array to check
				originalContentToCheck = sheetOrig[indexOrigX + str(indexOrigY)].value

				if changedContentToCheck == originalContentToCheck:
					match = True
					break;
			if match == True:
				break;

		# If there's no match, then stylize the cell
		if not match:
			print "Found cell without a match!"
			sheetChanged[indexChangedX + str(indexChangedY)].font = newFont

# Save the changed document
wb.save('styled_ss.xlsx')
print "Saved changes."